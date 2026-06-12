import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from compare_files import auto_compare, parse_key_spec, write_excel_report


class CompareFilesTests(unittest.TestCase):
    def test_parse_key_spec(self):
        self.assertEqual(parse_key_spec("1+5"), (0, 4))
        self.assertEqual(parse_key_spec(None), None)

    def test_compare_csv_with_key_and_excel_report(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            left = tmp_path / "left.csv"
            right = tmp_path / "right.csv"
            report = tmp_path / "report.xlsx"

            with left.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerows(
                    [
                        ["id", "name", "city"],
                        ["1", "Alice", "Rome"],
                        ["2", "Bob", "Milan"],
                    ]
                )

            with right.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerows(
                    [
                        ["id", "name", "city"],
                        ["1", "Alice", "Turin"],
                        ["3", "Carla", "Naples"],
                    ]
                )

            result = auto_compare(left, right, "1")
            write_excel_report(result, report)

            self.assertEqual(result.summary["changed"], 1)
            self.assertEqual(result.summary["added"], 1)
            self.assertEqual(result.summary["removed"], 1)

            workbook = load_workbook(report)
            sheet = workbook["Differences"]
            rows = list(sheet.iter_rows(values_only=True))
            workbook.close()

            self.assertIn(("CHANGED", "1", "city", "Rome", "Turin"), rows)
            self.assertIn(("REMOVED", "2", "id", "2", None), rows)
            self.assertIn(("ADDED", "3", "name", None, "Carla"), rows)

    def test_compare_text_files(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            left = tmp_path / "left.txt"
            right = tmp_path / "right.txt"

            left.write_text("alpha\nbeta\ngamma\n", encoding="utf-8")
            right.write_text("alpha\nbeta 2\ndelta\n", encoding="utf-8")

            result = auto_compare(left, right)

            self.assertEqual(result.mode, "text")
            self.assertEqual(result.summary["changed"], 2)
            self.assertEqual(result.differences[0]["status"], "CHANGED")


if __name__ == "__main__":
    unittest.main()
