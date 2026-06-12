from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Sequence

from docx import Document
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from pypdf import PdfReader

TABULAR_EXTENSIONS = {".csv", ".xlsx", ".xlsm"}
STATUS_FILLS = {
    "ADDED": PatternFill(fill_type="solid", fgColor="C6EFCE"),
    "REMOVED": PatternFill(fill_type="solid", fgColor="FFC7CE"),
    "CHANGED": PatternFill(fill_type="solid", fgColor="FFEB9C"),
}


@dataclass
class ComparisonResult:
    mode: str
    differences: list[dict[str, str | int]]
    summary: dict[str, str | int]


class ComparisonError(ValueError):
    pass


def parse_key_spec(key_spec: str | None) -> tuple[int, ...] | None:
    if not key_spec:
        return None

    indexes: list[int] = []
    for chunk in key_spec.split("+"):
        chunk = chunk.strip()
        if not chunk or not chunk.isdigit():
            raise ComparisonError(
                "La chiave deve essere nel formato '1+5' con indici numerici positivi."
            )
        index = int(chunk)
        if index <= 0:
            raise ComparisonError("Gli indici della chiave devono partire da 1.")
        indexes.append(index - 1)
    return tuple(indexes)


def read_csv_rows(path: Path) -> list[list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [[str(cell).strip() for cell in row] for row in csv.reader(handle)]


def read_xlsx_rows(path: Path) -> list[list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(["" if cell is None else str(cell).strip() for cell in row])
    workbook.close()
    return rows


def read_docx_lines(path: Path) -> list[str]:
    document = Document(path)
    return [paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip()]


def read_pdf_lines(path: Path) -> list[str]:
    reader = PdfReader(str(path))
    lines: list[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        lines.extend(line.strip() for line in text.splitlines() if line.strip())
    return lines


def read_doc_lines(path: Path) -> list[str]:
    raise ComparisonError(
        "I file .doc legacy non sono supportati direttamente. Converti il file in .docx oppure .pdf e riprova."
    )


def read_text_lines(path: Path) -> list[str]:
    extension = path.suffix.lower()
    if extension == ".txt":
        with path.open("r", encoding="utf-8-sig") as handle:
            return [line.rstrip("\n") for line in handle]
    if extension == ".csv":
        return ["\t".join(row) for row in read_csv_rows(path)]
    if extension in {".xlsx", ".xlsm"}:
        return ["\t".join(row) for row in read_xlsx_rows(path)]
    if extension == ".pdf":
        return read_pdf_lines(path)
    if extension == ".docx":
        return read_docx_lines(path)
    if extension == ".doc":
        return read_doc_lines(path)

    try:
        with path.open("r", encoding="utf-8-sig") as handle:
            return [line.rstrip("\n") for line in handle]
    except UnicodeDecodeError as exc:
        raise ComparisonError(f"Formato non supportato per la lettura testuale: {path.suffix}") from exc


def read_tabular_rows(path: Path) -> list[list[str]]:
    extension = path.suffix.lower()
    if extension == ".csv":
        return read_csv_rows(path)
    if extension in {".xlsx", ".xlsm"}:
        return read_xlsx_rows(path)
    raise ComparisonError(f"Formato tabellare non supportato: {path.suffix}")


def split_header(rows1: list[list[str]], rows2: list[list[str]]) -> tuple[list[str], list[list[str]], list[list[str]]]:
    if rows1 and rows2 and len(rows1[0]) == len(rows2[0]) and rows1[0] == rows2[0]:
        header = [cell or f"Colonna {index + 1}" for index, cell in enumerate(rows1[0])]
        return header, rows1[1:], rows2[1:]

    width = max((len(row) for row in rows1 + rows2), default=0)
    header = [f"Colonna {index + 1}" for index in range(width)]
    return header, rows1, rows2


def pad_row(row: Sequence[str], width: int) -> list[str]:
    return list(row) + [""] * (width - len(row))


def build_record_key(row: Sequence[str], row_number: int, key_indexes: tuple[int, ...] | None) -> str:
    if key_indexes is None:
        return str(row_number)
    try:
        return " | ".join(row[index] for index in key_indexes)
    except IndexError as exc:
        raise ComparisonError("La chiave fa riferimento a colonne non presenti nel file.") from exc


def rows_to_mapping(rows: list[list[str]], key_indexes: tuple[int, ...] | None, width: int) -> dict[str, list[str]]:
    mapping: dict[str, list[str]] = {}
    for offset, raw_row in enumerate(rows, start=1):
        row = pad_row(raw_row, width)
        record_key = build_record_key(row, offset, key_indexes)
        if record_key in mapping:
            raise ComparisonError(f"Chiave duplicata trovata nel file: {record_key}")
        mapping[record_key] = row
    return mapping


def compare_tabular_files(file1: Path, file2: Path, key_indexes: tuple[int, ...] | None) -> ComparisonResult:
    rows1 = read_tabular_rows(file1)
    rows2 = read_tabular_rows(file2)
    header, data1, data2 = split_header(rows1, rows2)
    width = len(header)
    mapping1 = rows_to_mapping(data1, key_indexes, width)
    mapping2 = rows_to_mapping(data2, key_indexes, width)

    differences: list[dict[str, str | int]] = []
    added = removed = changed = 0

    for record_key in sorted(set(mapping1) | set(mapping2)):
        row1 = mapping1.get(record_key)
        row2 = mapping2.get(record_key)

        if row1 is None and row2 is not None:
            added += 1
            for column_name, value2 in zip(header, row2):
                differences.append(
                    {
                        "status": "ADDED",
                        "record_key": record_key,
                        "column": column_name,
                        "file1": "",
                        "file2": value2,
                    }
                )
            continue

        if row2 is None and row1 is not None:
            removed += 1
            for column_name, value1 in zip(header, row1):
                differences.append(
                    {
                        "status": "REMOVED",
                        "record_key": record_key,
                        "column": column_name,
                        "file1": value1,
                        "file2": "",
                    }
                )
            continue

        assert row1 is not None and row2 is not None
        row_changed = False
        for column_name, value1, value2 in zip(header, row1, row2):
            if value1 != value2:
                row_changed = True
                differences.append(
                    {
                        "status": "CHANGED",
                        "record_key": record_key,
                        "column": column_name,
                        "file1": value1,
                        "file2": value2,
                    }
                )
        if row_changed:
            changed += 1

    return ComparisonResult(
        mode="tabular",
        differences=differences,
        summary={
            "file1": str(file1),
            "file2": str(file2),
            "key": "auto-riga" if key_indexes is None else "+".join(str(index + 1) for index in key_indexes),
            "records_file1": len(data1),
            "records_file2": len(data2),
            "added": added,
            "removed": removed,
            "changed": changed,
        },
    )


def compare_text_files(file1: Path, file2: Path) -> ComparisonResult:
    lines1 = read_text_lines(file1)
    lines2 = read_text_lines(file2)
    matcher = SequenceMatcher(a=lines1, b=lines2)
    differences: list[dict[str, str | int]] = []
    added = removed = changed = 0

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            continue
        if tag == "delete":
            for offset, line in enumerate(lines1[i1:i2], start=i1 + 1):
                removed += 1
                differences.append(
                    {"status": "REMOVED", "line": offset, "file1": line, "file2": ""}
                )
            continue
        if tag == "insert":
            for offset, line in enumerate(lines2[j1:j2], start=j1 + 1):
                added += 1
                differences.append(
                    {"status": "ADDED", "line": offset, "file1": "", "file2": line}
                )
            continue

        left = lines1[i1:i2]
        right = lines2[j1:j2]
        max_len = max(len(left), len(right))
        for index in range(max_len):
            changed += 1
            differences.append(
                {
                    "status": "CHANGED",
                    "line": max(i1, j1) + index + 1,
                    "file1": left[index] if index < len(left) else "",
                    "file2": right[index] if index < len(right) else "",
                }
            )

    return ComparisonResult(
        mode="text",
        differences=differences,
        summary={
            "file1": str(file1),
            "file2": str(file2),
            "lines_file1": len(lines1),
            "lines_file2": len(lines2),
            "added": added,
            "removed": removed,
            "changed": changed,
        },
    )


def auto_compare(file1: Path, file2: Path, key_spec: str | None = None) -> ComparisonResult:
    key_indexes = parse_key_spec(key_spec)
    both_tabular = file1.suffix.lower() in TABULAR_EXTENSIONS and file2.suffix.lower() in TABULAR_EXTENSIONS
    if both_tabular:
        return compare_tabular_files(file1, file2, key_indexes)
    if key_indexes is not None:
        raise ComparisonError("Il parametro --key è disponibile solo per confronti tabellari CSV/XLSX.")
    return compare_text_files(file1, file2)


def write_excel_report(result: ComparisonResult, output_path: Path) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Campo", "Valore"])
    for field, value in result.summary.items():
        summary_sheet.append([field, value])
    summary_sheet.freeze_panes = "A2"

    details_sheet = workbook.create_sheet("Differences")
    if result.mode == "tabular":
        headers = ["Status", "Record Key", "Column", "File 1", "File 2"]
        details_sheet.append(headers)
        for difference in result.differences:
            details_sheet.append(
                [
                    difference["status"],
                    difference["record_key"],
                    difference["column"],
                    difference["file1"],
                    difference["file2"],
                ]
            )
    else:
        headers = ["Status", "Line", "File 1", "File 2"]
        details_sheet.append(headers)
        for difference in result.differences:
            details_sheet.append(
                [
                    difference["status"],
                    difference["line"],
                    difference["file1"],
                    difference["file2"],
                ]
            )

    details_sheet.freeze_panes = "A2"
    details_sheet.auto_filter.ref = details_sheet.dimensions
    for row in details_sheet.iter_rows(min_row=2):
        status = row[0].value
        fill = STATUS_FILLS.get(status)
        if fill:
            for cell in row:
                cell.fill = fill

    for sheet in (summary_sheet, details_sheet):
        for column_cells in sheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 60)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Confronta due file e genera un report Excel con le differenze."
    )
    parser.add_argument("file1", type=Path, help="Primo file da confrontare")
    parser.add_argument("file2", type=Path, help="Secondo file da confrontare")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("differenze.xlsx"),
        help="Percorso del file Excel di output",
    )
    parser.add_argument(
        "-k",
        "--key",
        dest="key_spec",
        help="Chiave record per confronti tabellari, nel formato 1+5",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    if not args.file1.exists() or not args.file2.exists():
        parser.error("Entrambi i file di input devono esistere.")

    try:
        result = auto_compare(args.file1, args.file2, args.key_spec)
        write_excel_report(result, args.output)
    except ComparisonError as exc:
        parser.error(str(exc))

    print(f"Report generato: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
